import time
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.action_chains import ActionChains
from selenium.common.exceptions import NoSuchElementException
import openpyxl
from openpyxl.styles import Alignment
import tkinter as tk
from tkinter import filedialog
import os

class AutomacaoExati:
    def __init__(self):
        self.driver = None
        self.wait_time = 5  # Default wait time in seconds

    def selecionar_planilha(self):
        """Solicita ao usuário que selecione a planilha Excel"""
        root = tk.Tk()
        root.withdraw()  # Esconde a janela principal
        
        # Configuração da caixa de diálogo
        file_path = filedialog.askopenfilename(
            title="Selecione a planilha Excel",
            filetypes=[("Arquivos Excel", "*.xlsx;*.xls"), ("Todos os arquivos", "*.*")],
            initialdir=os.path.expanduser("~")
        )
        
        if not file_path:
            print("Nenhum arquivo selecionado. O programa será encerrado.")
            return None
        
        return file_path

    def inicializar_chrome(self):
        """1. Inicialização do Navegador"""
        options = webdriver.ChromeOptions()
        options.add_argument("--start-maximized")
        self.driver = webdriver.Chrome(options=options)
    
    def login_exati(self, excel_file):
        """2. Login no site"""
        try:
            wb = openpyxl.load_workbook(excel_file)
            ws = wb["EXATI"]
            
            url = ws['B1'].value
            username = ws['B2'].value
            password = ws['B3'].value
            
            if not all([url, username, password]):
                raise ValueError("Dados de login incompletos na planilha EXATI")
            
            self.driver.get(url)
            time.sleep(2)
            
            self.driver.find_element(By.ID, "userInfo.username").send_keys(username)
            self.driver.find_element(By.ID, "userInfo.password").send_keys(password)
            
            # Clica no botão "Acessar"
            self.driver.find_element(By.XPATH, "//button[.//span[contains(.,'Acessar')]]").click()
            time.sleep(5)
            
        except Exception as e:
            print(f"Erro durante o login: {str(e)}")
            raise

    # ... (mantenha o restante dos métodos iguais ao código anterior)

    def executar_processo_completo(self):
        """Executa todo o processo"""
        try:
            # Solicita ao usuário que selecione a planilha
            excel_file = self.selecionar_planilha()
            if not excel_file:
                return
            
            print(f"Planilha selecionada: {excel_file}")
            
            self.inicializar_chrome()
            self.login_exati(excel_file)
            self.configuracoes_pos_login()
            self.buscar_plaquetas(excel_file)
            
        except Exception as e:
            print(f"Erro durante a execução: {str(e)}")
            if self.driver:
                self.driver.quit()


# Exemplo de uso
if __name__ == "__main__":
    automacao = AutomacaoExati()
    automacao.executar_processo_completo()
